"""
build_laudo.py
Altcom 365 v2 — geração dos laudos Excel (cliente + interno Altcom).

Suporta dois formatos:
  • Novo (UPPERCASE) — Relatório Milvus Completo com 15+ colunas obrigatórias
  • Antigo (mixed case) — export Analítico por cliente (retrocompatibilidade)

Fluxo recomendado para novo formato:
  1. calcular_alertas(df_uppercase, versao_ref)   → df com _alerta_* e _uso_pct
  2. normalize_df(df_com_alertas)                 → renomeia para colunas do engine
  3. build_laudo_cliente(df, output, cliente)
  4. build_relatorio_interno(df, output, cliente, versao_ref)
"""
import os, sys, re
from datetime import datetime
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from engine_altcom365 import (classify, device_type, parse_uso, parse_storage,
                               parse_ram, is_win_old, BADGE_COLORS)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, WHITE, ZEBRA, BORDER_C = "0D1B2A", "FFFFFF", "F5F7FA", "D0D5DD"
CYAN = "00C8D4"

# -- Mapeamento novo formato → colunas do engine -------------------------------
COL_MAP_NOVO = {
    'NOME DO DISPOSITIVO':         'Nome do dispositivo',
    'TIPO DO DISPOSITIVO':         'Tipo de dispositivo',
    'SISTEMA OPERACIONAL':         'Sistema operacional',
    'PROCESSADOR':                 'Processador',
    'NÚCLEOS DO PROCESSADOR':      'Núcleos do processador',
    'MEMÓRIA RAM TOTAL':           'Memória RAM total',
    'ARMAZENAMENTO INTERNO TOTAL': 'Armazenamento total',
    'NOME FANTASIA DO CLIENTE':    'Cliente',
    'APELIDO':                     'Apelido',
    'USUÁRIO LOGADO':              'Usuário logado',
    'DATA DE ATUALIZAÇÃO':         'Data de atualização',
    'VERSÃO DO CLIENT':            'Versão do client',
}

# -- Frases que NÃO aparecem no laudo do cliente -------------------------------
_DESC_REMOVE = [
    "Necessário fazer upgrade para Windows 11.",
    "Manutenção preventiva de armazenamento recomendada.",
    "Armazenamento em nível crítico — limpeza urgente necessária.",
]
_SUG_REMOVE = {"Upgrade Windows 11 Pro"}  # storage warnings kept for client
_WIN_PRICE  = "R$ 145,00"

# Cores de alerta interno
ALERT_COLORS = {
    'armazenamento': ("FFE7E7", "C0392B"),
    'windows':       ("FFF2CC", "D4AC0D"),
    'sem_contato':   ("D4E8F8", "2874A6"),
    'milvus':        ("ECEFF1", "5D6D7E"),
    'troca':         ("F9E6E6", "922B21"),
}


# ==============================================================================
# NORMALIZAÇÃO
# ==============================================================================

def is_new_format(df):
    """Retorna True se o df vem do Relatório Milvus Completo (colunas UPPERCASE)."""
    return 'NOME DO DISPOSITIVO' in df.columns


def normalize_df(df):
    """
    Converte df do novo formato (UPPERCASE) para colunas compatíveis com o engine.
    Se já estiver no formato antigo, retorna sem alterações.
    Preserva colunas _uso_pct e _alerta_* geradas por alertas_internos.
    """
    if not is_new_format(df):
        return df  # antigo — engine já funciona

    df = df.copy()

    # -- Calcula Armazenamento utilizado % -------------------------------------
    if '_uso_pct' in df.columns:
        # Já computado por calcular_alertas → converte para string
        df['Armazenamento utilizado'] = df['_uso_pct'].apply(
            lambda u: f"{u:.2f}%" if (u is not None and not pd.isna(u)) else "NaN%"
        )
    else:
        def _calc_uso(row):
            def _gb(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return None
                s = str(v).replace(' GB','').replace('GB','').replace(',','.').strip()
                try: return float(s)
                except: return None
            total = _gb(row.get('ARMAZENAMENTO INTERNO TOTAL'))
            if not total or total <= 0:
                return "NaN%"
            util = _gb(row.get('ARMAZENAMENTO INTERNO UTILIZADO'))
            if util is not None:
                return f"{util/total*100:.2f}%"
            disp = _gb(row.get('ARMAZENAMENTO INTERNO DISPONÍVEL'))
            if disp is not None:
                return f"{(total-disp)/total*100:.2f}%"
            return "NaN%"
        df['Armazenamento utilizado'] = df.apply(_calc_uso, axis=1)

    # -- Renomeia colunas ------------------------------------------------------
    df = df.rename(columns=COL_MAP_NOVO)

    # -- Limpa strings em colunas de texto opcionais ---------------------------
    for col in ('Apelido', 'Usuário logado'):
        if col in df.columns:
            df[col] = (df[col].astype(str)
                       .replace({'nan': '', 'None': '', 'Não possui': '', 'NaN': ''}))

    return df


# ==============================================================================
# ESTILOS OPENPYXL
# ==============================================================================

def brd():
    s = Side(style='thin', color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def badge_style(classif_base, badge_text=None):
    bg, fg = BADGE_COLORS.get(classif_base, ("FFFFFF", "000000"))
    return PatternFill("solid", fgColor=bg), Font(name='Arial', bold=True, size=8, color=fg)

def hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill      = PatternFill("solid", fgColor=CYAN)
    cell.font      = Font(name='Arial', bold=True, size=9, color=NAVY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = brd()

def dat(ws, r, c, v, z=False, ha='left', wrap=True, size=8):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill      = PatternFill("solid", fgColor=ZEBRA if z else WHITE)
    cell.font      = Font(name='Arial', size=size)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border    = brd()

def title_strip(ws, row, text, span, h=34):
    ws.row_dimensions[row].height = h
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)


# ==============================================================================
# HELPERS DE LIMPEZA (laudo do cliente)
# ==============================================================================

def _clean_descritivo(desc):
    result = str(desc)
    for phrase in _DESC_REMOVE:
        result = result.replace(phrase, '')
    result = re.sub(r'\s+', ' ', result).strip()
    return result or "Configuração adequada."

def _clean_sugestao(sug):
    if not sug or str(sug) in ('NA', 'nan'):
        return "NA"
    lines = [l.strip() for l in str(sug).split('\n') if l.strip() not in _SUG_REMOVE]
    return '\n'.join(lines) if lines else "NA"

def _clean_precos(prec):
    if not prec or str(prec) in ('NA', 'nan'):
        return "NA"
    parts = [p.strip() for p in str(prec).split(' | ') if p.strip() != _WIN_PRICE]
    return ' | '.join(parts) if parts else "NA"

def _motivos_upgrade(row):
    """Retorna lista dos motivos que geraram sufixo '- Upgrade'.
    Valores possíveis: 'so', 'ram', 'ssd'."""
    motivos = []
    if is_win_old(str(row.get('Sistema operacional', ''))):
        motivos.append('so')
    if parse_ram(row.get('Memória RAM total', 0)) < 8:
        motivos.append('ram')
    if parse_storage(row.get('Armazenamento total', 0)) < 200:
        motivos.append('ssd')
    return motivos


def _badge_cliente(row):
    """Badge para o laudo cliente.
    '- Upgrade' por SO apenas + uso <= 70% → classificação base (sem sufixo).
    '- Upgrade' por SO apenas + uso > 70%  → classif + '- Man. Prev.' (uso alto é visível ao cliente).
    '- Upgrade' por RAM/SSD (com ou sem SO) → mantém '- Upgrade'.
    '- Man. Prev.' e CRÍTICO               → mantidos sem alteração.
    """
    badge        = str(row.get('Badge', ''))
    classif_base = str(row.get('Classificação', badge))
    if '- Upgrade' in badge:
        motivos = _motivos_upgrade(row)
        if motivos == ['so']:
            # Win 10 é o único motivo — verifica uso de armazenamento
            uso = parse_uso(row.get('Armazenamento utilizado', 'NaN%'))
            if uso is not None and uso > 70:
                return classif_base + ' - Man. Prev.'
            return classif_base
    return badge

def _parse_data_at(val):
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return val
    if hasattr(val, 'date'):
        return pd.Timestamp(val)
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'nat', ''):
        return None
    for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return pd.to_datetime(s, format=fmt)
        except Exception:
            pass
    try:
        return pd.to_datetime(s, dayfirst=True)
    except Exception:
        return None

def _uso_display(row):
    """Retorna string 'XX.X%' para exibição, preferindo _uso_pct pré-computado."""
    u = row.get('_uso_pct')
    if u is not None and not (isinstance(u, float) and pd.isna(u)):
        return f"{float(u):.1f}%"
    uso = parse_uso(row.get('Armazenamento utilizado', 'NaN%'))
    return f"{uso:.1f}%" if uso is not None else "N/D"


# ==============================================================================
# LAUDO DO CLIENTE (3 abas, limpo — 14 colunas)
# ==============================================================================

def build_laudo_cliente(df, output_path, cliente_nome=None):
    """
    Gera o laudo de eficiência para o cliente (Excel 3 abas).

    Parâmetros
    ----------
    df           : DataFrame já normalizado (após normalize_df)
    output_path  : caminho de saída .xlsx
    cliente_nome : string com nome do cliente (se None, lê de df['Cliente'])
    """
    results = df.apply(classify, axis=1)
    df_out  = pd.concat([df.reset_index(drop=True), results.reset_index(drop=True)], axis=1)

    if cliente_nome is None:
        cliente_nome = (str(df_out['Cliente'].iloc[0]).strip()
                        if 'Cliente' in df_out.columns else "Cliente")
    hoje  = pd.Timestamp.today().strftime('%d/%m/%y')
    total = len(df_out)

    wb = Workbook()

    # -- ABA 1: LAUDO ---------------------------------------------------------
    ws = wb.active; ws.title = "Laudo"
    NCOLS = 14

    title_strip(ws, 1, f"LAUDO DE EFICIÊNCIA TÉCNICA  |  {cliente_nome.upper()}", NCOLS)
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    ws['A2'] = f"Metodologia Altcom 365 — Avaliação do Parque de Estações de Trabalho — {hoje}"
    ws['A2'].font      = Font(name='Arial', size=9, color="888888")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 5

    HEADERS = ['Tipo', 'Dispositivo', 'Apelido', 'Usuário Logado', 'S.O.',
               'Processador', 'Núcleos', 'RAM', 'Armazenamento', 'Uso %',
               'Classificação', 'Descritivo', 'Durabilidade', 'Sugestão']
    ws.row_dimensions[4].height = 22
    for ci, h in enumerate(HEADERS, 1):
        hdr(ws, 4, ci, h)

    for i, (_, row) in enumerate(df_out.iterrows()):
        r = i + 5
        ws.row_dimensions[r].height = 46
        z = i % 2 == 0

        tipo_label   = ("Desktop"
                        if device_type(str(row.get('Tipo de dispositivo', ''))) == 'desktop'
                        else "Notebook")
        classif_base = str(row.get('Classificação', ''))
        badge_c      = _badge_cliente(row)  # remove sufixo se upgrade for só SO
        uso_s        = _uso_display(row)
        st           = parse_storage(row.get('Armazenamento total', 0))
        st_s         = f"{st:.0f} GB SSD" if st > 0 else "N/D"
        desc_c       = _clean_descritivo(str(row.get('Descritivo', '')))
        sug_c        = _clean_sugestao(str(row.get('Sugestão', row.get('Sugestao', ''))))
        dur_raw      = str(row.get('Durabilidade estimada', ''))
        apelido      = str(row.get('Apelido', '')) or '—'
        usuario      = str(row.get('Usuário logado', '')) or '—'

        vals = [
            (tipo_label,                            'center'),
            (str(row.get('Nome do dispositivo', '')), 'left'),
            (apelido,                                'left'),
            (usuario,                                'left'),
            (str(row.get('Sistema operacional', '')), 'left'),
            (str(row.get('Processador', '')),          'left'),
            (str(row.get('Núcleos do processador', '')), 'center'),
            (str(row.get('Memória RAM total', '')),    'center'),
            (st_s,    'center'),
            (uso_s,   'center'),
            (badge_c, 'center'),
            (desc_c,  'left'),
            (dur_raw, 'center'),
            (sug_c,   'left'),
        ]

        for ci, (v, ha) in enumerate(vals, 1):
            if ci == 11:  # Classificação — badge colorido
                c = ws.cell(row=r, column=ci, value=v)
                fill, font = badge_style(classif_base)
                c.fill = fill; c.font = font
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = brd()
            else:
                dat(ws, r, ci, v, z=z, ha=ha)

    for i, w in enumerate([10, 20, 16, 16, 26, 36, 8, 8, 14, 7, 20, 50, 12, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

    # -- ABA 2: RESUMO EXECUTIVO -----------------------------------------------
    ws2 = wb.create_sheet("Resumo Executivo")
    title_strip(ws2, 1, "RESUMO EXECUTIVO", 5)
    ws2.row_dimensions[2].height = 18
    ws2.merge_cells('A2:E2')
    ws2['A2'] = f"Total de dispositivos avaliados: {total} | Cliente: {cliente_nome} | Data: {hoje}"
    ws2['A2'].font      = Font(name='Arial', size=10, color="444444")
    ws2['A2'].alignment = Alignment(indent=1)
    ws2.row_dimensions[3].height = 5
    for ci, h in enumerate(['Classificação', 'Qtd', '%', 'Durabilidade ref.', 'Status'], 1):
        hdr(ws2, 4, ci, h)
    ws2.row_dimensions[4].height = 22

    classif_col = 'Classificação'
    order_real  = ["EXCELENTE", "ÓTIMO", "BOM", "SATISFATÓRIO", "CRÍTICO"]
    dur_ref_r   = {"EXCELENTE": "2029/2030", "ÓTIMO": "2028/2029", "BOM": "2027/2028",
                   "SATISFATÓRIO": "2026/2027", "CRÍTICO": "Troca"}
    status_r    = {"EXCELENTE": "Hardware ideal, alta longevidade",
                   "ÓTIMO":     "Hardware adequado, operação estável",
                   "BOM":       "Funcional, ajustes pontuais necessários",
                   "SATISFATÓRIO": "Limite operacional, planejar renovação",
                   "CRÍTICO":   "Substituição imediata recomendada"}

    for ri, cat in enumerate(order_real, 5):
        qtd = int((df_out[classif_col] == cat).sum())
        ws2.row_dimensions[ri].height = 24
        bg, fg = BADGE_COLORS[cat]
        c = ws2.cell(row=ri, column=1, value=cat)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name='Arial', bold=True, size=9, color=fg)
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = brd()
        for ci, v in enumerate([qtd, f"{qtd/total*100:.0f}%", dur_ref_r[cat], status_r[cat]], 2):
            cc = ws2.cell(row=ri, column=ci, value=v)
            cc.font      = Font(name='Arial', size=9)
            cc.alignment = Alignment(horizontal='center' if ci < 5 else 'left', vertical='center')
            cc.border    = brd()

    for ci, w in enumerate([16, 8, 8, 16, 38], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # -- ABA 3: LEGENDA --------------------------------------------------------
    ws3 = wb.create_sheet("Legenda")
    title_strip(ws3, 1, "LEGENDA — METODOLOGIA ALTCOM 365", 3)
    ws3.row_dimensions[2].height = 5
    legend_rows = [
        ("Classificação", "Critério base", "Sufixo / Ação"),
        ("EXCELENTE",    "CPU topo de linha + 16 GB RAM + SSD ≥ 480 GB + Win 11", "—"),
        ("ÓTIMO",        "CPU moderna + requisitos mínimos atendidos", "—"),
        ("BOM",          "CPU intermediária superior + requisitos mínimos", "—"),
        ("SATISFATÓRIO", "CPU intermediária inferior ou requisitos parciais", "—"),
        ("CRÍTICO",      "CPU obsoleta — substituição necessária", "—"),
        ("",             "", ""),
        ("Sufixo",       "Condição",                       "Ação recomendada"),
        ("- Upgrade",    "RAM < 8 GB ou SSD < 200 GB",    "Upgrade de componente"),
    ]
    for ri, (a, b, c_val) in enumerate(legend_rows, 3):
        ws3.row_dimensions[ri].height = 22
        if ri in (3, 11):
            for ci, v in enumerate([a, b, c_val], 1): hdr(ws3, ri, ci, v)
        elif a in BADGE_COLORS:
            bg, fg = BADGE_COLORS[a]
            cell = ws3.cell(row=ri, column=1, value=a)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name='Arial', bold=True, size=9, color=fg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = brd()
            for ci, v in enumerate([b, c_val], 2):
                dat(ws3, ri, ci, v, ha='left')
        else:
            for ci, v in enumerate([a, b, c_val], 1):
                dat(ws3, ri, ci, v, ha='left' if ci > 1 else 'center')
    for ci, w in enumerate([18, 58, 28], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)


# ==============================================================================
# RELATÓRIO INTERNO ALTCOM (1 aba, 16 colunas)
# ==============================================================================

def build_relatorio_interno(df, output_path, cliente_nome=None, versao_ref=None):
    """
    Gera o relatório interno para a equipe Altcom (Excel 1 aba).

    Inclui apenas dispositivos com pelo menos 1 alerta ativo.
    Dispositivos CRÍTICO aparecem com 'Laudado para Troca' em todas as colunas de alerta.

    Parâmetros
    ----------
    df          : DataFrame normalizado (após normalize_df); preferencialmente com _alerta_*
    output_path : caminho de saída .xlsx
    cliente_nome: string com nome do cliente
    versao_ref  : string da versão de referência do agente Milvus (opcional)
    """
    hoje     = pd.Timestamp.today()
    hoje_str = hoje.strftime('%d/%m/%y %H:%M')

    if cliente_nome is None:
        cliente_nome = (str(df['Cliente'].iloc[0]).strip()
                        if 'Cliente' in df.columns else "Cliente")

    HAS_DATA_AT   = 'Data de atualização' in df.columns
    HAS_VERSAO    = 'Versão do client'    in df.columns
    HAS_ALERTAS   = '_tem_alerta'          in df.columns

    results  = df.apply(classify, axis=1)
    df_out   = pd.concat([df.reset_index(drop=True), results.reset_index(drop=True)], axis=1)

    # -- Calcula alertas inline se não vieram pré-computados ------------------
    if not HAS_ALERTAS:
        hoje_ts = hoje
        def _inline_alerts(row):
            uso = parse_uso(row.get('Armazenamento utilizado', 'NaN%'))
            al_arm = f"Preventiva — uso {uso:.1f}%" if (uso and uso > 70) else ""
            so = str(row.get('Sistema operacional', '')).lower()
            al_win = "Upgrade para Win 11" if 'windows 10' in so else ""
            al_sc = ""
            if HAS_DATA_AT:
                dt = _parse_data_at(row.get('Data de atualização'))
                if dt and (hoje_ts - dt).days > 20:
                    al_sc = f"{int((hoje_ts-dt).days)} dias sem contato — validar com cliente"
            al_ml = ""
            if versao_ref and HAS_VERSAO:
                v = str(row.get('Versão do client', '')).strip()
                if v and v not in ('', 'nan', 'Não possui') and v != str(versao_ref):
                    al_ml = f"Desatualizada ({v}) — atualizar"
            tem = bool(al_arm or al_win or al_sc or al_ml)
            uso_pct = uso  # raw float
            return pd.Series({
                '_uso_pct': uso_pct,
                '_alerta_armazenamento': al_arm,
                '_alerta_windows':       al_win,
                '_alerta_sem_contato':   al_sc,
                '_alerta_milvus':        al_ml,
                '_tem_alerta':           tem,
            })
        inline = df_out.apply(_inline_alerts, axis=1)
        df_out = pd.concat([df_out, inline], axis=1)

    # -- Filtra apenas dispositivos com alerta ---------------------------------
    alert_df = df_out[df_out['_tem_alerta']].copy()

    # -- Cabeçalho Excel -------------------------------------------------------
    HEADERS = ['Dispositivo', 'Apelido', 'Usuário Logado', 'Tipo', 'Cliente',
               'S.O.', 'Processador', 'RAM', 'Armazenamento', 'Uso %',
               'Data Atualização', 'Versão Agente',
               'Alerta Armazenamento', 'Alerta Windows',
               'Alerta Sem Contato', 'Alerta Agente Milvus']
    NCOLS = len(HEADERS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas Internos"

    title_strip(ws, 1,
                f"RELATÓRIO INTERNO — ALERTAS  |  {cliente_nome.upper()}", NCOLS)
    ws.row_dimensions[2].height = 18
    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    ws['A2'] = f"Dispositivos com ações operacionais pendentes — gerado em {hoje_str}"
    ws['A2'].font      = Font(name='Arial', size=9, color="888888")
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[3].height = 5

    ws.row_dimensions[4].height = 22
    for ci, h in enumerate(HEADERS, 1):
        hdr(ws, 4, ci, h)

    if alert_df.empty:
        ws.merge_cells(f'A5:{get_column_letter(NCOLS)}5')
        c = ws.cell(row=5, column=1, value="Nenhum alerta identificado no parque.")
        c.font      = Font(name='Arial', size=10, italic=True, color="888888")
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[5].height = 28
        wb.save(output_path)
        return

    for i, (_, row) in enumerate(alert_df.iterrows()):
        r  = i + 5
        ws.row_dimensions[r].height = 40
        z  = i % 2 == 0

        tipo_label   = ("Desktop"
                        if device_type(str(row.get('Tipo de dispositivo', ''))) == 'desktop'
                        else "Notebook")
        classif_base = str(row.get('Classificação', ''))
        badge_text   = str(row.get('Badge', classif_base))
        uso_s        = _uso_display(row)
        st           = parse_storage(row.get('Armazenamento total', 0))
        st_s         = f"{st:.0f} GB SSD" if st > 0 else "N/D"
        apelido      = str(row.get('Apelido', '')) or '—'
        usuario      = str(row.get('Usuário logado', '')) or '—'
        cliente_val  = str(row.get('Cliente', cliente_nome)) or cliente_nome
        data_at_val  = str(row.get('Data de atualização', 'N/D'))
        versao_val   = str(row.get('Versão do client', 'N/D'))

        # Colunas básicas (1-12)
        base_vals = [
            (str(row.get('Nome do dispositivo', '')), 'left'),
            (apelido,      'left'),
            (usuario,      'left'),
            (tipo_label,   'center'),
            (cliente_val,  'left'),
            (str(row.get('Sistema operacional', '')), 'left'),
            (str(row.get('Processador', '')),          'left'),
            (str(row.get('Memória RAM total', '')),    'center'),
            (st_s,         'center'),
            (uso_s,        'center'),
            (data_at_val if HAS_DATA_AT else 'N/D', 'center'),
            (versao_val  if HAS_VERSAO  else 'N/D', 'center'),
        ]
        ci = 1
        for v, ha in base_vals:
            dat(ws, r, ci, v, z=z, ha=ha)
            ci += 1

        # Colunas de alerta (13-16)
        if classif_base == 'CRÍTICO':
            # Dispositivo laudado para troca — sem procedimentos específicos
            bg_t, fc_t = ALERT_COLORS['troca']
            for col_ci in range(ci, ci + 4):
                is_first = col_ci == ci
                c = ws.cell(row=r, column=col_ci,
                            value="Laudado para Troca" if is_first else "")
                c.fill      = PatternFill("solid", fgColor=bg_t)
                c.font      = Font(name='Arial', size=8, bold=is_first, color=fc_t)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border    = brd()
        else:
            alert_vals = [
                (str(row.get('_alerta_armazenamento', '')), 'armazenamento'),
                (str(row.get('_alerta_windows', '')),       'windows'),
                (str(row.get('_alerta_sem_contato', '')),   'sem_contato'),
                (str(row.get('_alerta_milvus', '')),        'milvus'),
            ]
            for v, color_key in alert_vals:
                bg, fc = ALERT_COLORS[color_key]
                c = ws.cell(row=r, column=ci, value=v if v else "")
                if v:
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.font = Font(name='Arial', size=8, bold=True, color=fc)
                else:
                    c.fill = PatternFill("solid", fgColor=ZEBRA if z else WHITE)
                    c.font = Font(name='Arial', size=8, color="CCCCCC")
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border    = brd()
                ci += 1

    # -- Larguras --------------------------------------------------------------
    widths = [22, 16, 16, 9, 20, 22, 34, 8, 14, 7, 18, 16, 22, 18, 28, 26]
    for ci, w in enumerate(widths[:NCOLS], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A5'

    # -- Rodape -----------------------------------------------------------------
    last_r = (len(alert_df) + 5) if not alert_df.empty else 6
    ws.merge_cells(f'A{last_r}:{get_column_letter(NCOLS)}{last_r}')
    wb.save(output_path)
